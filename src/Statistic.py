import pandas as pd
import os
import decimal
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from Visual import visualize
from Run import Res_Dir, Summarize_DIR, reward_funs

decimal.getcontext().rounding = "ROUND_HALF_UP"
res_csv = os.path.join('results', 'summarize', 'add_result.csv')

saveFileName = 'CRL-TCP.xlsx'


def analysis_csv(df, relist, exlist):
    lista = []

    for column, env in enumerate(sorted(df['env'].unique())):
        env1 = env.title()
        env1 = env1.replace("_", " ")
        if env1 == "Iofrol":
            env1 = "IOF/ROL"
        slist = [env1]

        for title in ['napfd', 'recall', 'ttf', 'durations']:
            for re in relist:
                tdf = df[df["env"].isin([env]) &
                         df["rewardfun"].isin([re])
                         ]

                slist.append(float(decimal.Decimal(str(tdf[title]).split()[1]).quantize(
                    decimal.Decimal("0.0000"))))

        lista.append(slist)

    slist = ['Average']
    rownums = len(lista)
    colnums = len(lista[0])

    for i in range(1, colnums):
        cc = 0
        for j in range(0, rownums):
            cc += lista[j][i]
        slist.append(float(decimal.Decimal(cc / rownums).quantize(
            decimal.Decimal("0.0000"))))
    lista.append(slist)

    cdf = pd.DataFrame(lista, columns=exlist)

    # Save to excel
    cdf.to_excel(saveFileName, index=False)


def mark_csv(interval):
    wb = load_workbook(filename=saveFileName, data_only=True)
    work = wb[wb.sheetnames[0]]

    all = []
    for row in work.rows:
        lista = []
        for c in row:
            lista.append(c.value)
        all.append(lista)

    print(len(all[2]))
    base = interval

    maxindex = []

    for i in range(1, len(all)):
        for j in range(1, 3):
            index = all[i].index(max(all[i][base * (j - 1) + 1:base * j + 1]))
            if index >= 26:
                s = "A" + chr(index + 65 - 26) + str(i + 1)
            else:
                s = chr(index + 65) + str(i + 1)
            maxindex.append(s)

        for j in range(3, 5):
            index = all[i].index(min(all[i][base * (j - 1) + 1:base * j + 1]))
            if index >= 26:
                s = "A" + chr(index + 65 - 26) + str(i + 1)
            else:
                s = chr(index + 65) + str(i + 1)
            maxindex.append(s)

    fill = PatternFill('solid', fgColor='FFFF00')
    font = Font(bold=True)
    for i in maxindex:
        work[i].font = font
        work[i].fill = fill

    wb.close()
    wb.save(saveFileName)


if __name__ == '__main__':
    if not os.path.exists(res_csv):
        print(f'{res_csv} Does not exist, being regenerated')
        visualize(name='add', Res_Dir=Res_Dir, Summarize_DIR=Summarize_DIR)
    df = pd.read_csv(res_csv)
    print(df.head())

    relist = list(reward_funs.keys())
    exlist = ["Dataset"] + relist * 4

    if len(relist) == 1:
        exlist = ["Dataset", "NAPFD", "Recall", "TTF", "Durations"]
        analysis_csv(df, relist, exlist)
    else:
        analysis_csv(df, relist, exlist)
        mark_csv(len(relist))
